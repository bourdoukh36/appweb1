import os
import tkinter as tk
from tkinter import ttk, messagebox
from PIL import Image, ImageTk  # pip install pillow
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import openpyxl
from datetime import datetime
import subprocess
import sys


# ---------- CRÉATION .VENV (À EXÉCUTER UNE FOIS) ----------
def create_venv():
    """Crée l'environnement virtuel .venv si absent"""
    venv_path = os.path.join(os.path.dirname(__file__), '.venv')
    if not os.path.exists(venv_path):
        subprocess.check_call([sys.executable, '-m', 'venv', '.venv'])
        print("✅ .venv créé avec succès!")
        print("Installez les dépendances: .venv\\Scripts\\activate && pip install gspread oauth2client openpyxl pillow")
    else:
        print("✅ .venv existe déjà.")


create_venv()

# ---------- CONFIG LOGO ----------
LOGO_PATH = r"C:\Users\hp\PycharmProjects\operations\logo.png"  # Adaptez extension si .jpg/.ico
LOGO_SIZE = (120, 120)  # Taille affichée (largeur, hauteur)

# ---------- CONFIG GOOGLE SHEETS ----------
SCOPE = ["https://spreadsheets.google.com/feeds",
         'https://www.googleapis.com/auth/spreadsheets',
         "https://www.googleapis.com/auth/drive.file",
         "https://www.googleapis.com/auth/drive"]

creds = ServiceAccountCredentials.from_json_keyfile_name(
    r'C:\Users\hp\PycharmProjects\suivi des opérations\credentials.json', SCOPE
)
client = gspread.authorize(creds)
SHEET_NAME = "suivi des opérations"

# ---------- DONNÉES FIXES ----------
SERRES = ['B', 'C', 'D', 'E', 'F', 'G', 'H']
DELTAS = [str(i) for i in range(1, 33)]
CULTURES = ['tomate', 'pastèque', 'poivron', 'concombre', 'laitue', 'ciboulette', 'courgette', 'herbes aromatiques']
TRAITEMENTS = ['fongicide', 'insecticide', 'acaricide', 'insecticide/acaricide', 'raticide', 'bio-stimulant',
               'désinfectant', 'engrais foliaire']
SOLUTIONS_IRRI = ['AB', 'CD', 'M', 'Urée', 'enracineur', 'désinfectant']
ECS = ['1.6', '1.8', '2', '2.5', '3', '3.5', '4']

EXCEL_PRODUITS = "produits.xlsx"

# ---------- CRÉATION AUTOMATIQUE PRODUITS.XLSX ----------
if not os.path.exists(EXCEL_PRODUITS):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Produits"
    ws.append(["Designation", "Dose", "Cible"])
    wb.save(EXCEL_PRODUITS)
    print(f"{EXCEL_PRODUITS} créé avec succès.")


# ---------- FONCTIONS UTILITAIRES ----------
def ajouter_produit(designation, dose, cible):
    """Ajoute un produit dans produits.xlsx"""
    wb = openpyxl.load_workbook(EXCEL_PRODUITS)
    ws = wb.active
    ws.append([designation, dose, cible])
    wb.save(EXCEL_PRODUITS)
    print(f"Produit '{designation}' enregistré dans {EXCEL_PRODUITS}.")


def charger_produits():
    """Charge tous les produits depuis produits.xlsx"""
    wb = openpyxl.load_workbook(EXCEL_PRODUITS)
    ws = wb.active
    produits = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and len(row) >= 3:
            produits.append((row[0], row[1], row[2]))
    return produits


# ---------- FORMULAIRE AJOUT PRODUIT ----------
def form_ajout_produit(app):
    fen = tk.Toplevel(app.root)
    fen.title("Ajouter Produit")
    fen.geometry("300x250")
    # Ajout logo sur formulaire
    try:
        logo_img = Image.open(LOGO_PATH)
        logo_img = logo_img.resize(LOGO_SIZE, Image.Resampling.LANCZOS)
        logo_photo = ImageTk.PhotoImage(logo_img)
        logo_label = tk.Label(fen, image=logo_photo, bg='white')
        logo_label.image = logo_photo  # Référence
        logo_label.pack(pady=10)
    except:
        tk.Label(fen, text="🌱", font=('Arial', 40)).pack(pady=10)

    tk.Label(fen, text="Designation").pack(pady=5)
    des_var = tk.StringVar()
    tk.Entry(fen, textvariable=des_var, width=30).pack(pady=5)

    tk.Label(fen, text="Dose").pack(pady=5)
    dose_var = tk.StringVar()
    tk.Entry(fen, textvariable=dose_var, width=30).pack(pady=5)

    tk.Label(fen, text="Cible").pack(pady=5)
    cible_var = tk.StringVar()
    tk.Entry(fen, textvariable=cible_var, width=30).pack(pady=5)

    def enregistrer():
        if not all([des_var.get(), dose_var.get(), cible_var.get()]):
            messagebox.showerror("Erreur", "Remplissez tous les champs")
            return
        ajouter_produit(des_var.get(), dose_var.get(), cible_var.get())
        messagebox.showinfo("Succès", f"Produit {des_var.get()} ajouté")
        fen.destroy()

    tk.Button(fen, text="Ajouter", command=enregistrer, bg='green', fg='white', width=15).pack(pady=20)


# ---------- CLASSE PRINCIPALE ----------
class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Suivi Opérations Pépinière - Multi-Delta & Multi-Produits")
        self.root.geometry("900x700")
        self.sheet = None
        self.selected_deltas = []
        self.selected_produits_details = []
        self.produits = []
        self.client = client
        self.logo_img = None
        self.logo_photo = None
        self.create_widgets()

    def load_logo(self):
        """Charge et redimensionne le logo"""
        try:
            self.logo_img = Image.open(LOGO_PATH)
            self.logo_img = self.logo_img.resize(LOGO_SIZE, Image.Resampling.LANCZOS)
            self.logo_photo = ImageTk.PhotoImage(self.logo_img)
            return True
        except Exception as e:
            print(f"Erreur logo: {e}. Vérifiez {LOGO_PATH}")
            return False

    def create_widgets(self):
        # ---------- FRAME LOGO EN-TÊTE ----------
        header_frame = tk.Frame(self.root, bg='white', relief='raised', bd=2)
        header_frame.grid(row=0, column=0, columnspan=6, sticky='ew', padx=10, pady=10)

        if self.load_logo():
            logo_label = tk.Label(header_frame, image=self.logo_photo, bg='white')
            logo_label.image = self.logo_photo  # Référence pour éviter GC
            logo_label.pack(side='left', padx=10)
        else:
            logo_label = tk.Label(header_frame, text="🌱", font=('Arial', 50), bg='lightgreen')
            logo_label.pack(side='left', padx=10)

        title_label = tk.Label(header_frame, text="Suivi Opérations Pépinière",
                               font=('Arial', 20, 'bold'), bg='white', fg='darkgreen')
        title_label.pack(side='left', padx=20)

        # Icône fenêtre (optionnel, si logo.ico)
        try:
            icon_img = Image.open(LOGO_PATH)
            icon_photo = ImageTk.PhotoImage(icon_img.resize((32, 32)))
            self.root.iconphoto(False, icon_photo)
        except:
            pass

        # Séparateur
        ttk.Separator(self.root, orient='horizontal').grid(row=1, column=0, columnspan=6, sticky='ew', pady=10)

        # Serre (row 2 maintenant)
        tk.Label(self.root, text="Serre:").grid(row=2, column=0, sticky='w', padx=10, pady=5)
        self.serre_cb = ttk.Combobox(self.root, values=SERRES, state='readonly', width=10)
        self.serre_cb.grid(row=2, column=1, padx=5)
        self.serre_cb.bind('<<ComboboxSelected>>', self.on_serre_change)

        # Multi-Delta (row 2)
        tk.Label(self.root, text="Deltas (Multi):").grid(row=2, column=2, sticky='w', padx=10, pady=5)
        delta_frame = tk.Frame(self.root)
        delta_frame.grid(row=2, column=3, columnspan=2, padx=5, sticky='w')
        self.delta_lb = tk.Listbox(delta_frame, selectmode='multiple', height=4, width=15, exportselection=0)
        self.delta_lb.pack(side='left')
        for delta in DELTAS:
            self.delta_lb.insert(tk.END, delta)
        tk.Button(delta_frame, text="Valider", command=self.valider_deltas, width=12).pack(side='left', padx=5)

        self.deltas_label = tk.Label(self.root, text="Sélectionnés: Aucun", fg='blue')
        self.deltas_label.grid(row=3, column=3, columnspan=2, sticky='w', padx=10)

        # Culture (row 3)
        tk.Label(self.root, text="Culture:").grid(row=3, column=0, sticky='w', padx=10, pady=5)
        self.culture_cb = ttk.Combobox(self.root, values=CULTURES, state='readonly', width=20)
        self.culture_cb.grid(row=3, column=1, columnspan=2, padx=5, pady=5)

        # Opération (row 4)
        tk.Label(self.root, text="Opération:").grid(row=4, column=0, sticky='w', padx=10, pady=5)
        self.operation_cb = ttk.Combobox(self.root, values=['traitement', 'irrigation'], state='readonly', width=15)
        self.operation_cb.grid(row=4, column=1, padx=5, pady=5)
        self.operation_cb.bind('<<ComboboxSelected>>', self.on_operation_change)

        # Frame détails (row 5)
        self.details_frame = tk.Frame(self.root)
        self.details_frame.grid(row=5, column=0, columnspan=6, pady=10, padx=10, sticky='ew')

        # Boutons (row 6)
        btn_frame = tk.Frame(self.root)
        btn_frame.grid(row=6, column=0, columnspan=6, pady=20)
        tk.Button(btn_frame, text="💾 Enregistrer", command=self.enregistrer, bg='green', fg='white', width=12,
                  font=('Arial', 10, 'bold')).pack(side='left', padx=10)
        tk.Button(btn_frame, text="➕ Ajouter Produit", command=lambda: form_ajout_produit(self), bg='blue', fg='white',
                  width=15).pack(side='left', padx=10)
        tk.Button(btn_frame, text="🔄 Actualiser Produits", command=self.update_produits, bg='orange', fg='white',
                  width=15).pack(side='left', padx=10)

        self.root.grid_columnconfigure(5, weight=1)
        self.root.grid_rowconfigure(5, weight=1)

    # ---------- METHODES (INCHANGÉES) ----------
    def on_serre_change(self, event):
        self.selected_deltas = []
        self.deltas_label.config(text="Sélectionnés: Aucun")
        self.delta_lb.selection_clear(0, tk.END)

    def valider_deltas(self):
        selection = self.delta_lb.curselection()
        self.selected_deltas = [DELTAS[i] for i in selection]
        self.deltas_label.config(
            text=f"Sélectionnés: {', '.join(self.selected_deltas) if self.selected_deltas else 'Aucun'}")

    def update_produits(self):
        self.produits = charger_produits()
        if hasattr(self, 'produit_lb'):
            self.produit_lb.delete(0, tk.END)
            for p in self.produits:
                self.produit_lb.insert(tk.END, p[0])

    def on_operation_change(self, event):
        for widget in self.details_frame.winfo_children():
            widget.destroy()
        op = self.operation_cb.get()
        row = 0
        if op == 'traitement':
            tk.Label(self.details_frame, text="Traitement:").grid(row=row, column=0, sticky='w')
            self.trait_cb = ttk.Combobox(self.details_frame, values=TRAITEMENTS, state='readonly', width=20)
            self.trait_cb.grid(row=row, column=1, padx=5)
            row += 1

            tk.Label(self.details_frame, text="Produits (Multi):").grid(row=row, column=0, sticky='w')
            prod_frame = tk.Frame(self.details_frame)
            prod_frame.grid(row=row, column=1, columnspan=2, sticky='w', pady=5)
            self.produit_lb = tk.Listbox(prod_frame, selectmode='multiple', height=6, width=25, exportselection=0)
            self.produit_lb.pack(side='left')
            self.update_produits()
            tk.Button(prod_frame, text="Valider", command=self.valider_produits, width=12).pack(side='left', padx=5)
            row += 1

            self.produits_details_label = tk.Label(self.details_frame, text="Détails: Aucun", fg='green',
                                                   justify='left')
            self.produits_details_label.grid(row=row, column=0, columnspan=3, sticky='w', pady=5)

        elif op == 'irrigation':
            tk.Label(self.details_frame, text="Solution:").grid(row=row, column=0, sticky='w')
            self.solution_cb = ttk.Combobox(self.details_frame, values=SOLUTIONS_IRRI, state='readonly', width=15)
            self.solution_cb.grid(row=row, column=1, padx=5)
            row += 1
            tk.Label(self.details_frame, text="EC:").grid(row=row, column=0, sticky='w')
            self.ec_cb = ttk.Combobox(self.details_frame, values=ECS, state='readonly', width=15)
            self.ec_cb.grid(row=row, column=1, padx=5)

    def valider_produits(self):
        selection = self.produit_lb.curselection()
        self.selected_produits_details = []
        for i in selection:
            prod = self.produits[i]
            self.selected_produits_details.append(f"{prod[0]} {prod[1]} {prod[2]}")
        details_text = '; '.join(self.selected_produits_details) if self.selected_produits_details else "Aucun"
        self.produits_details_label.config(text=f"Détails: {details_text}")

    def get_or_create_sheet(self, serre_delta):
        sh = self.client.open(SHEET_NAME)
        feuille_nom = serre_delta
        try:
            self.sheet = sh.worksheet(feuille_nom)
        except gspread.WorksheetNotFound:
            self.sheet = sh.add_worksheet(title=feuille_nom, rows=1000, cols=20)
            headers = ['Date', 'Serre', 'Delta', 'Culture', 'Operation', 'Details']
            self.sheet.append_row(headers)

    def enregistrer(self):
        if not self.serre_cb.get() or not self.selected_deltas or not self.culture_cb.get() or not self.operation_cb.get():
            messagebox.showerror("Erreur", "Remplissez tous les champs principaux (Deltas obligatoires)")
            return

        serre = self.serre_cb.get()
        culture = self.culture_cb.get()
        operation = self.operation_cb.get()
        date = datetime.now().strftime("%Y-%m-%d %H:%M")
        deltas_str = ', '.join(self.selected_deltas)

        if operation == 'traitement':
            trait = self.trait_cb.get() or ''
            details = f"{trait} - {self.produits_details_label.cget('text').replace('Détails: ', '')}" if self.selected_produits_details else "Aucun produit"
        else:
            details = f"{self.solution_cb.get() or ''} EC{self.ec_cb.get() or ''}"

        for delta in self.selected_deltas:
            self.get_or_create_sheet(f"{serre}{delta}")
            row = [date, serre, delta, culture, operation, details]
            self.sheet.append_row(row)

        messagebox.showinfo("Succès", f"Enregistré dans {serre}{deltas_str}\nDétails: {details}")
        self.selected_deltas = []
        self.selected_produits_details = []
        self.deltas_label.config(text="Sélectionnés: Aucun")
        self.delta_lb.selection_clear(0, tk.END)
        if hasattr(self, 'produit_lb'):
            self.produit_lb.selection_clear(0, tk.END)


# ---------- LANCEMENT ----------
if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()
