import streamlit as st
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from datetime import datetime
import openpyxl
import os
import pandas as pd
from PIL import Image

# ---------- CONFIG ----------
SCOPE = [
    "https://spreadsheets.google.com/feeds",
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive.file",
    "https://www.googleapis.com/auth/drive",
]


@st.cache_resource
def init_google_sheets():
    credentials_dict = st.secrets["google"]
    creds = ServiceAccountCredentials.from_json_keyfile_dict(credentials_dict, SCOPE)
    return gspread.authorize(creds)


client = init_google_sheets()
SHEET_NAME = "suivi des opérations"

SERRES = ['B', 'C', 'D', 'E', 'F', 'G', 'H']
DELTAS = [str(i) for i in range(1, 33)]
CULTURES = ['tomate', 'pastèque', 'poivron', 'concombre', 'laitue', 'ciboulette', 'courgette', 'herbes aromatiques']
TRAITEMENTS = ['fongicide', 'insecticide', 'acaricide', 'insecticide/acaricide', 'raticide',
               'bio-stimulant', 'désinfectant', 'engrais foliaire']
SOLUTIONS_IRRI = ['AB', 'CD', 'M', 'Urée', 'enracineur', 'désinfectant']
ECS = ['1.6', '1.8', '2', '2.5', '3', '3.5', '4']

EXCEL_PRODUITS = "produits.xlsx"

# ---------- SESSION STATE SIMPLIFIÉ ----------
if 'success_message' not in st.session_state:
    st.session_state.success_message = False
if 'form_data' not in st.session_state:
    st.session_state.form_data = {}


# ---------- CRÉATION AUTOMATIQUE PRODUITS.XLSX ----------
def create_produits_excel():
    if not os.path.exists(EXCEL_PRODUITS):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Produits"
        ws.append(["Designation", "Dose", "Cible"])
        wb.save(EXCEL_PRODUITS)


create_produits_excel()


# ---------- FONCTIONS ----------
@st.cache_data
def charger_produits():
    try:
        wb = openpyxl.load_workbook(EXCEL_PRODUITS)
        ws = wb.active
        produits = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= 3:
                produits.append({
                    'nom': str(row[0]).strip(),
                    'dose': str(row[1]).strip(),
                    'cible': str(row[2]).strip(),
                    'details': f"{row[0]} {row[1]} {row[2]}"
                })
        return produits
    except:
        return []


def ajouter_produit(designation, dose, cible):
    wb = openpyxl.load_workbook(EXCEL_PRODUITS)
    ws = wb.active
    ws.append([designation, dose, cible])
    wb.save(EXCEL_PRODUITS)
    st.cache_data.clear()


def get_or_create_sheet(client, serre_delta):
    sh = client.open(SHEET_NAME)
    try:
        return sh.worksheet(serre_delta)
    except gspread.WorksheetNotFound:
        sheet = sh.add_worksheet(title=serre_delta, rows=1000, cols=20)
        headers = ['Date', 'Serre', 'Delta', 'Culture', 'Operation', 'Details']
        sheet.append_row(headers)
        return sheet


def get_details_produits(selected_noms, produits):
    details_produits = []
    for nom in selected_noms:
        for p in produits:
            if p['nom'] == nom:
                details_produits.append(p['details'])
                break
    return details_produits


def reset_form():
    """Réinitialise complètement le formulaire"""
    st.session_state.success_message = True
    st.session_state.form_data = {}
    st.rerun()


# ---------- INTERFACE STREAMLIT ----------
st.set_page_config(
    page_title="Suivi Opérations Pépinière",
    page_icon="🌱",
    layout="wide"
)

# Sidebar
with st.sidebar:
    st.title("🌱 Pépinière")

    st.subheader("📦 Produits")
    produits = charger_produits()

    with st.form("ajout_produit"):
        des = st.text_input("**Designation**", placeholder="ex: Amistar")
        dose = st.text_input("**Dose**", placeholder="ex: 2ml/L")
        cible = st.text_input("**Cible**", placeholder="ex: pucerons")
        submitted = st.form_submit_button("➕ **Ajouter**")

    if submitted and all([des, dose, cible]):
        ajouter_produit(des, dose, cible)
        st.success(f"✅ **{des}** ajouté!")
        st.rerun()

    if produits:
        st.markdown("**Produits disponibles:**")
        for i, p in enumerate(produits, 1):
            st.write(f"{i}. **{p['nom']}** ({p['dose']})")

# Message de succès
if st.session_state.success_message:
    st.success("🎉 **ENREGISTRÉ AVEC SUCCÈS!** Tous les champs ont été réinitialisés.")
    st.balloons()
    if st.button("🔄 **Nouvelle Opération**"):
        st.session_state.success_message = False
        st.session_state.form_data = {}
        st.rerun()
    st.markdown("---")

# Formulaire principal (UNIQUES SANS KEY CONFLICTS)
st.title("📊 Suivi Opérations Pépinière")

col1, col2, col3 = st.columns(3)
with col1:
    serre = st.selectbox("**Serre:**", SERRES)
with col2:
    selected_deltas = st.multiselect("**Deltas:**", DELTAS, max_selections=10)
with col3:
    culture = st.selectbox("**Culture:**", CULTURES)

operation = st.selectbox("**Opération:**", ['traitement', 'irrigation'])

# Champs conditionnels
selected_traitements = []
selected_noms = []
solution = ""
ec = ""

if operation == 'traitement':
    col_t1, col_t2 = st.columns(2)
    with col_t1:
        selected_traitements = st.multiselect("**Traitements:**", TRAITEMENTS, max_selections=4)
    with col_t2:
        if produits:
            noms_produits = [p['nom'] for p in produits]
            selected_noms = st.multiselect("**Produits:**", noms_produits, max_selections=8)

elif operation == 'irrigation':
    col_i1, col_i2 = st.columns(2)
    with col_i1:
        solution = st.selectbox("**Solution:**", SOLUTIONS_IRRI)
    with col_i2:
        ec = st.selectbox("**EC:**", ECS)

# Aperçu
with st.expander("👀 **Aperçu**", expanded=True):
    if operation == 'traitement':
        details = "; ".join(selected_traitements) if selected_traitements else "Aucun"
        if selected_noms:
            details_produits = get_details_produits(selected_noms, produits)
            if details_produits:
                details += f" - {'; '.join(details_produits)}"
    else:
        details = f"{solution} EC{ec}" if solution and ec else "Incomplete"

    st.info(f"""
    **Date:** {datetime.now().strftime("%Y-%m-%d %H:%M")}
    **Serre:** {serre}  
    **Deltas:** {', '.join(selected_deltas) if selected_deltas else 'Aucun'}  
    **Culture:** {culture}  
    **Opération:** {operation}  
    **Détails:** {details}
    """)

# BOUTON ENREGISTRER SIMPLE
if st.button("💾 **ENREGISTRER**", type="primary", use_container_width=True, disabled=st.session_state.success_message):
    # Validation
    if not all([serre, selected_deltas, culture]):
        st.error("❌ **Serre, Deltas et Culture obligatoires!**")
    elif operation == 'traitement' and not selected_traitements:
        st.error("❌ **Sélectionnez au moins 1 traitement!**")
    elif operation == 'irrigation' and not (solution and ec):
        st.error("❌ **Solution et EC obligatoires!**")
    else:
        # ENREGISTREMENT DIRECT
        date = datetime.now().strftime("%Y-%m-%d %H:%M")
        success_count = 0

        for delta in selected_deltas:
            try:
                sheet = get_or_create_sheet(client, f"{serre}{delta}")
                row = [date, serre, delta, culture, operation, details]
                sheet.append_row(row)
                success_count += 1
            except Exception as e:
                st.error(f"❌ Delta {delta}: {e}")

        if success_count > 0:
            st.session_state.success_message = True
            reset_form()
        else:
            st.error("❌ Erreur d'enregistrement!")

# Historique
if st.checkbox("📋 **Historique**") and serre and selected_deltas:
    try:
        if len(selected_deltas) == 1:
            sh = client.open(SHEET_NAME)
            feuille = sh.worksheet(f"{serre}{selected_deltas[0]}")
            data = feuille.get_all_values()
            if len(data) > 1:
                df = pd.DataFrame(data[1:], columns=data[0])
                st.dataframe(df.tail(15), use_container_width=True, height=400)
        else:
            st.warning("⚠️ Sélectionnez 1 seul delta pour l'historique")
    except Exception as e:
        st.error(f"❌ Google Sheets: {e}")

st.markdown("---")
st.markdown("🌱 **Suivi Pépinière - Simple & Efficace**")


